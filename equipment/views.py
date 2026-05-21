from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Equipment, EquipmentType, WriteOffRequest
from .forms import EquipmentForm
from core.models import Room, ChangeLog


@login_required
def equipment_list(request):
    qs = Equipment.objects.select_related('room__floor__building', 'equipment_type').all()
    search = request.GET.get('search', '')
    condition = request.GET.get('condition', '')
    eq_type = request.GET.get('type', '')
    floor_id = request.GET.get('floor', '')
    room_id = request.GET.get('room', '')
    is_present = request.GET.get('is_present', '')

    if request.user.is_keeper:
        qs = qs.filter(room__in=request.user.rooms_kept.all())

    if search:
        qs = qs.filter(Q(reg_number__icontains=search) | Q(brand__icontains=search) | Q(model__icontains=search))
    if condition:
        qs = qs.filter(condition=condition)
    if eq_type:
        qs = qs.filter(equipment_type_id=eq_type)
    if floor_id:
        qs = qs.filter(room__floor_id=floor_id)
    if room_id:
        qs = qs.filter(room_id=room_id)
    if is_present:
        qs = qs.filter(is_present=(is_present == '1'))

    from core.models import Floor
    ctx = {
        'equipment': qs,
        'equipment_types': EquipmentType.objects.all(),
        'floors': Floor.objects.select_related('building').all(),
        'rooms': Room.objects.all(),
        'conditions': Equipment.CONDITION_CHOICES,
        'filters': {'search': search, 'condition': condition, 'type': eq_type, 'floor': floor_id, 'room': room_id, 'is_present': is_present},
    }
    return render(request, 'equipment/list.html', ctx)


@login_required
def equipment_detail(request, pk):
    item = get_object_or_404(Equipment, pk=pk)
    logs = ChangeLog.objects.filter(model_name='Equipment', object_id=pk).order_by('-created_at')[:20]
    from requests_app.models import RepairRequest
    ctx = {'item': item, 'logs': logs, 'repair_requests': RepairRequest.objects.filter(equipment=item).order_by('-created_at')}
    return render(request, 'equipment/detail.html', ctx)


@login_required
def equipment_add(request):
    form = EquipmentForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        item = form.save(commit=False)
        item.created_by = request.user
        item.save()
        ChangeLog.objects.create(user=request.user, action=f'Добавлено оборудование: {item}', model_name='Equipment', object_id=item.id)
        messages.success(request, 'Оборудование добавлено')
        return redirect('equipment_detail', pk=item.pk)
    return render(request, 'equipment/form.html', {'form': form, 'title': 'Добавить оборудование'})


@login_required
def equipment_edit(request, pk):
    item = get_object_or_404(Equipment, pk=pk)
    old_data = {'condition': item.condition, 'cost': str(item.cost), 'is_present': item.is_present}
    form = EquipmentForm(request.POST or None, request.FILES or None, instance=item)
    if request.method == 'POST' and form.is_valid():
        item = form.save()
        ChangeLog.objects.create(user=request.user, action=f'Изменено оборудование: {item}', model_name='Equipment', object_id=item.id, old_value=old_data)
        messages.success(request, 'Данные обновлены')
        return redirect('equipment_detail', pk=item.pk)
    return render(request, 'equipment/form.html', {'form': form, 'title': 'Редактировать оборудование', 'item': item})


@login_required
def equipment_writeoff(request, pk):
    if not (request.user.is_admin or request.user.is_ahc):
        return redirect('dashboard')
    item = get_object_or_404(Equipment, pk=pk)
    if request.method == 'POST':
        confirm_word = request.POST.get('confirm_word', '').strip().upper()
        reason = request.POST.get('reason', '')
        if confirm_word == 'СПИСАТЬ':
            WriteOffRequest.objects.create(equipment=item, reason=reason, created_by=request.user, confirmed_by=request.user, confirmed_at=timezone.now())
            item.condition = Equipment.CONDITION_WRITTEN_OFF
            item.save()
            ChangeLog.objects.create(user=request.user, action=f'Оборудование списано: {item}', model_name='Equipment', object_id=item.id)
            messages.success(request, f'Оборудование {item} списано')
            return redirect('equipment_list')
        else:
            messages.error(request, 'Неверное слово подтверждения')
    return render(request, 'equipment/writeoff.html', {'item': item})


@login_required
def equipment_import(request):
    if not (request.user.is_admin or request.user.is_keeper):
        return redirect('dashboard')
    errors = []
    success_count = 0
    if request.method == 'POST' and request.FILES.get('file'):
        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active
        headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            try:
                data = dict(zip(headers, row))
                room = Room.objects.filter(number=str(data.get('Кабинет', ''))).first()
                eq_type, _ = EquipmentType.objects.get_or_create(name=str(data.get('Вид', 'Прочее')))
                Equipment.objects.update_or_create(
                    reg_number=str(data.get('Рег. номер', '')),
                    defaults={
                        'brand': str(data.get('Марка', '')),
                        'model': str(data.get('Модель', '')),
                        'specs': str(data.get('Характеристики', '')),
                        'quantity': int(data.get('Количество', 1) or 1),
                        'equipment_type': eq_type,
                        'room': room,
                        'cost': data.get('Стоимость') or None,
                        'condition': 'ok',
                        'created_by': request.user,
                    }
                )
                success_count += 1
            except Exception as e:
                errors.append(f'Строка {i}: {e}')
    return render(request, 'equipment/import.html', {'errors': errors, 'success_count': success_count})


@login_required
def equipment_export(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Оборудование'
    headers = ['Рег. номер', 'Марка', 'Модель', 'Вид', 'Кабинет', 'Этаж', 'Состояние', 'На месте', 'Количество', 'Дата внесения', 'Последняя проверка', 'Стоимость']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F3864')

    qs = Equipment.objects.select_related('room__floor', 'equipment_type').all()
    if request.user.is_keeper:
        qs = qs.filter(room__in=request.user.rooms_kept.all())

    for row_num, item in enumerate(qs, 2):
        ws.append([
            item.reg_number, item.brand, item.model,
            item.equipment_type.name if item.equipment_type else '',
            item.room.number if item.room else '',
            str(item.room.floor) if item.room else '',
            item.get_condition_display(),
            'Да' if item.is_present else 'Нет',
            item.quantity,
            str(item.date_added),
            str(item.last_check_date) if item.last_check_date else '',
            float(item.cost) if item.cost else '',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="equipment.xlsx"'
    wb.save(response)
    return response


@login_required
def download_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Шаблон'
    headers = ['Рег. номер', 'Марка', 'Модель', 'Характеристики', 'Вид', 'Кабинет', 'Количество', 'Стоимость']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2E75B6')
    ws.append(['ПК-001', 'Lenovo', 'ThinkCentre M720', 'Intel i5, 8GB RAM, 256 SSD', 'Компьютер', '101', 1, 35000])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="template.xlsx"'
    wb.save(response)
    return response
